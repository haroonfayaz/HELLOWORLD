import openpyxl as xl


wb2 = xl.load_workbook("amount9.xlsx")
sheet = wb2["Sheet1"]
cell = sheet.cell(2, 9)

sum1 = 0

for row in range (2, sheet.max_row + 1):
    amount = sheet.cell(row, 9)
    sum1 += amount.value
    sum1_cell = sheet.cell(row + 1 ,sheet.max_column + 2)
    sum1_cell.value =sum1

wb2.save("total91.xlsx")
xl.