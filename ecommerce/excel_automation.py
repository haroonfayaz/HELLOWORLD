import openpyxl as xl
from openpyxl.chart import BarChart, Reference


wb = xl.load_workbook("SALE ACCOUNTS.xlsx")
sheet = wb["Sheet1"]

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 4)
    calc_salary = cell.value * 0.1 * 60
    calc_salary_cell = sheet.cell(row, 9)
    calc_salary_cell.value = calc_salary

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          max_col=9,
          min_col=9)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,"j2",)

wb.remove("chart2.xlsx")